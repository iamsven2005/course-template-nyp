import zipfile              
from io import BytesIO        
from PIL import Image      
import base64                 
from openpyxl import load_workbook 

def extract_images_as_base64(file_stream, file_extension):
    """Extract images from an uploaded DOCX or XLSX file as base64-encoded strings."""
    images_base64 = []

    if file_extension == 'docx':
        with zipfile.ZipFile(file_stream, 'r') as z:
            for file_name in z.namelist():
                if file_name.startswith('word/media/'):
                    with z.open(file_name) as source_file:
                        image_data = source_file.read()
                        image_b64 = base64.b64encode(image_data).decode('utf-8')
                        images_base64.append((file_name, image_b64))
    elif file_extension == 'xlsx':
        workbook = load_workbook(file_stream)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for image in worksheet._images:
                image_data = image._data()
                image_b64 = base64.b64encode(image_data).decode('utf-8')
                images_base64.append((image.anchor._from, image_b64))

    return images_base64